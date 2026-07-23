"""Generate professional Excel reports for completed executions."""

from __future__ import annotations

from collections import Counter
from collections.abc import Iterable
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..models import ExecutionMode, ItemResult, ItemStatus


HEADER_FILL = PatternFill("solid", fgColor="8B0D17")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="E9EEF5")
THIN_GRAY = Side(style="thin", color="D1D5DB")
STATUS_FILLS = {
    ItemStatus.SUBMITTED.value: PatternFill("solid", fgColor="D1FAE5"),
    ItemStatus.SUBMITTED_UNCONFIRMED.value: PatternFill("solid", fgColor="FEF3C7"),
    ItemStatus.FILLED.value: PatternFill("solid", fgColor="DBEAFE"),
    ItemStatus.VALIDATED.value: PatternFill("solid", fgColor="E0E7FF"),
    ItemStatus.SKIPPED.value: PatternFill("solid", fgColor="F3F4F6"),
    ItemStatus.ERROR.value: PatternFill("solid", fgColor="FEE2E2"),
}


def _autosize(worksheet, max_width: int = 55) -> None:
    for index, column_cells in enumerate(worksheet.columns, start=1):
        width = 0
        for cell in column_cells:
            if cell.value is not None:
                width = max(width, len(str(cell.value)))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), max_width)


def build_result_workbook(
    results: Iterable[ItemResult],
    execution_id: str,
    mode: ExecutionMode,
    started_at: datetime,
    finished_at: datetime,
) -> bytes:
    result_list = list(results)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumo"
    data_sheet = workbook.create_sheet("Resultados")

    summary["A1"] = "AUTOMACAO DE ENCERRAMENTO DE PASTAS - GCPJ"
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = HEADER_FILL
    summary.merge_cells("A1:D1")
    summary["A3"] = "ID da execucao"
    summary["B3"] = execution_id
    summary["A4"] = "Modo"
    summary["B4"] = mode.label
    summary["A5"] = "Inicio"
    summary["B5"] = started_at
    summary["A6"] = "Fim"
    summary["B6"] = finished_at
    summary["A7"] = "Total de linhas"
    summary["B7"] = len(result_list)

    counts = Counter(result.status.value for result in result_list)
    summary["A9"] = "Status"
    summary["B9"] = "Quantidade"
    for cell in summary[9]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)
    for row_index, status in enumerate(ItemStatus, start=10):
        summary.cell(row=row_index, column=1, value=status.value)
        summary.cell(row=row_index, column=2, value=counts.get(status.value, 0))
        if status.value in STATUS_FILLS:
            summary.cell(row=row_index, column=1).fill = STATUS_FILLS[status.value]
            summary.cell(row=row_index, column=2).fill = STATUS_FILLS[status.value]

    if result_list:
        records = [result.to_record() for result in result_list]
        headers: list[str] = []
        for record in records:
            for key in record:
                if key not in headers:
                    headers.append(key)
        data_sheet.append(headers)
        for record in records:
            data_sheet.append([record.get(header, "") for header in headers])

        for cell in data_sheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        data_sheet.freeze_panes = "A2"
        data_sheet.auto_filter.ref = data_sheet.dimensions
        table = Table(displayName="TabelaResultados", ref=data_sheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        data_sheet.add_table(table)

        status_column = headers.index("_STATUS_AUTOMACAO") + 1
        for row in range(2, data_sheet.max_row + 1):
            status_value = str(data_sheet.cell(row=row, column=status_column).value or "")
            if status_value in STATUS_FILLS:
                data_sheet.cell(row=row, column=status_column).fill = STATUS_FILLS[status_value]

    for sheet in (summary, data_sheet):
        sheet.sheet_view.showGridLines = False
        _autosize(sheet)
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 34
    for row in summary.iter_rows(min_row=3, max_row=summary.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(bottom=THIN_GRAY)
    summary["B5"].number_format = "dd/mm/yyyy hh:mm:ss"
    summary["B6"].number_format = "dd/mm/yyyy hh:mm:ss"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_report(report_bytes: bytes, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(report_bytes)
    return output_path
