from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from gcpj_automation.models import ExecutionMode, ItemResult, ItemStatus
from gcpj_automation.services.report import build_result_workbook


def test_build_report_workbook():
    started = datetime(2026, 7, 22, 10, 0, 0)
    finished = datetime(2026, 7, 22, 10, 0, 3)
    result = ItemResult(
        row_number=2,
        npc="1600000001",
        ato="IMPROCEDENTE",
        gcpj_reason="IMPROCEDENCIA",
        status=ItemStatus.FILLED,
        message="Preenchido.",
        started_at=started,
        finished_at=finished,
        closure_date=date(2026, 7, 22),
        details="Encerramento conforme ATO.",
        original={"NPC": 1600000001, "ATO": "IMPROCEDENTE"},
    )
    payload = build_result_workbook(
        [result],
        execution_id="GCPJ_TESTE",
        mode=ExecutionMode.FILL_ONLY,
        started_at=started,
        finished_at=finished,
    )
    workbook = load_workbook(BytesIO(payload), data_only=False)
    assert workbook.sheetnames == ["Resumo", "Resultados"]
    assert workbook["Resumo"]["B3"].value == "GCPJ_TESTE"
    headers = [cell.value for cell in workbook["Resultados"][1]]
    assert "_STATUS_AUTOMACAO" in headers
    assert "_DATA_ENCERRAMENTO" in headers
    assert "_DETALHES_ENCERRAMENTO" in headers
