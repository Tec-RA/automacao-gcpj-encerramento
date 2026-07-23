from pathlib import Path

from openpyxl import load_workbook


def test_excel_template_contains_required_columns_and_guidance():
    template = Path(__file__).resolve().parents[1] / "templates" / "modelo_planilha_encerramento.xlsx"
    workbook = load_workbook(template)
    assert workbook.sheetnames == ["Encerramentos", "Instruções"]

    sheet = workbook["Encerramentos"]
    headers = [sheet.cell(1, column).value for column in range(1, sheet.max_column + 1)]
    assert "NPC" in headers
    assert "ATO" in headers
    assert "STATUS" in headers
    assert "CONTA PRA META" in headers
    assert sheet.freeze_panes == "A2"
    assert "TabelaEncerramentos" in sheet.tables
    assert len(sheet.data_validations.dataValidation) >= 2
